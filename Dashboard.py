# =========================
# DASHBOARD COMPLETO (RÁPIDO + FLUIDO + MAPA NÃO RECARREGA AO BAIXAR)
# - Leitura robusta (detecta mojibake e re-lê em UTF-8)
# - Cache agressivo
# - Mapa memoizado por assinatura (download não reconstrói)
# - Saúde:
#   * Corrige nomes estranhos (AmbulatÃ³rio etc.)
#   * Profissionais (STAFF_*) no Painel + Popups
#   * Remove o card "Unidades de Saúde (Total)" do Painel
#   * Mantém títulos: "Unidades por tipo" e "Profissionais"
# - Exportação XLSX estável:
#   * "Registro Atual" virou "Total"
#   * Coluna separada "Delta" com diferença e porcentagem (Cenário - Total)
#   * Cenário permanece numérico (bom para análise); Delta é texto: "+123 (+10,5%)"
# - ✅ TEMA LIGHT FORÇADO + HEADER BRANCO (ícones OK) SEM MUDAR MENU
# =========================

from pathlib import Path
import re
import hashlib
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st

import geopandas as gpd
import folium
from folium.plugins import MarkerCluster

try:
    from streamlit_folium import st_folium
except Exception:
    st.error("Instale streamlit-folium: pip install streamlit-folium")
    raise

# ---- Excel (exportação)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None
    Font = PatternFill = Alignment = Border = Side = None
    get_column_letter = None

# =========================
# CAMINHOS (DIRETO NA PASTA, SEM DISCO G:)
# =========================
APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "Dados"

MUNICIPIOS_DATA = {
    "Lajeado": {
        "empresas": str(DATA_DIR / "Lajeado" / "Empresas.xlsx"),
        "educacao": str(DATA_DIR / "Lajeado" / "Escolas.csv"),
        "saude": str(DATA_DIR / "Lajeado" / "Saúde.csv"),
    },
    "Porto Alegre": {
        "empresas": str(DATA_DIR / "Porto Alegre" / "Empresas.xlsx"),
        "educacao": str(DATA_DIR / "Porto Alegre" / "Escolas.csv"),
        "saude": str(DATA_DIR / "Porto Alegre" / "Saúde.csv"),
    },
    "Rio Grande": {
        "empresas": str(DATA_DIR / "Rio Grande" / "Empresas.xlsx"),
        "educacao": str(DATA_DIR / "Rio Grande" / "Escolas.csv"),
        "saude": str(DATA_DIR / "Rio Grande" / "Saúde.csv"),
    },
}

CENARIOS = {
    "Lajeado": {
        "Cenário 27m": str(DATA_DIR / "Lajeado" / "Mancha" / "27m00cm.shp"),
        "Cenário 30m": str(DATA_DIR / "Lajeado" / "Mancha" / "30m00cm.shp"),
    },
    "Porto Alegre": {
        "Cenário ADA": str(DATA_DIR / "Porto Alegre" / "Mancha" / "enchente_poa_intersects.shp"),
    },
    "Rio Grande": {
        "Cenário Setembro 2023": str(DATA_DIR / "Rio Grande" / "Mancha" / "CEN_SET2023.shp"),
        "Cenário Maio 2024": str(DATA_DIR / "Rio Grande" / "Mancha" / "CEN_MAI2024.shp"),
        "Cenário Maio 2024 + 50%": str(DATA_DIR / "Rio Grande" / "Mancha" / "CEN_MAI24_MAIS60CM.shp"),
    },
}

RS_CENTER = (-30.35, -53.35)
RS_ZOOM = 6.5

MUNICIPIO_VIEW = {
    "Lajeado": {"center": (-29.4585, -51.9953), "zoom": 12.5},
    "Porto Alegre": {"center": (-30.1051, -51.1500), "zoom": 11},
    "Rio Grande": {"center": (-32.0650, -52.1200), "zoom": 12.5},
}

PLACEHOLDER_EMP = "Selecione a Camada Empresas"
PLACEHOLDER_EDU = "Selecione a Camada Educação"
PLACEHOLDER_SAU = "Selecione a Camada Saúde"

# =========================
# SAÚDE - STAFFS
# =========================
STAFF_COLS = [
    "staff_acs_endemias",
    "staff_admin_gestao_apoio",
    "staff_diag_lab_imagem",
    "staff_enfermagem",
    "staff_farmacia",
    "staff_medicos",
    "staff_odontologia",
    "staff_outros",
    "staff_outros_superior_saude",
    "staff_servicos_gerais",
    "staff_transporte_urgencia",
]

STAFF_LABELS = {
    "staff_acs_endemias": "ACS/Endemias",
    "staff_admin_gestao_apoio": "Admin/Gestão/Apoio",
    "staff_diag_lab_imagem": "Diag/Lab/Imagem",
    "staff_enfermagem": "Enfermagem",
    "staff_farmacia": "Farmácia",
    "staff_medicos": "Médicos",
    "staff_odontologia": "Odontologia",
    "staff_outros": "Outros",
    "staff_outros_superior_saude": "Outros (Sup. Saúde)",
    "staff_servicos_gerais": "Serviços Gerais",
    "staff_transporte_urgencia": "Transp. Urgência",
}


def is_placeholder(val: str, placeholder: str) -> bool:
    if val is None:
        return True
    v = str(val).strip()
    return (v == "") or (v == placeholder)


# =========================
# UI / THEME + AJUSTES
# =========================
def inject_css():
    st.markdown(
        """
        <style>
        /* =========================================================
           ✅ FORÇAR LIGHT (SEM ALTERAR SEU MENU)
           - Deixa o app light mesmo com sistema dark
        ========================================================= */
        :root { color-scheme: light !important; }
        html, body, .stApp { background:#fff !important; color:#000 !important; }

        /* =========================================================
           ✅ SEU CSS ORIGINAL (MANTIDO)
        ========================================================= */
        .block-container { padding-top: 2.1rem !important; padding-bottom: 0.9rem !important; padding-left: 1.0rem !important; padding-right: 1.0rem !important; }
        h1 { margin:0 !important; line-height:1.10 !important; }

        section[data-testid="stSidebar"][aria-expanded="true"] + div div[data-testid="stAppViewContainer"] .main .block-container{
          padding-left: 1.0rem !important;
          padding-right: 1.0rem !important;
        }
        section[data-testid="stSidebar"][aria-expanded="true"]{
          width: 360px !important; min-width: 360px !important; max-width: 360px !important;
          background:#fff !important;
        }
        section[data-testid="stSidebar"][aria-expanded="false"]{
          width: 0px !important; min-width: 0px !important; max-width: 0px !important;
          overflow: hidden !important;
        }
        section[data-testid="stSidebar"] * { color:#000 !important; }
        section[data-testid="stSidebar"] div[data-testid="stSidebarContent"]{ padding-top: 0.2rem !important; }

        .block-container > div[data-testid="stHorizontalBlock"] > div:nth-child(2) > div[data-testid="stVerticalBlock"]{
          border: 2px solid #111 !important;
          border-radius: 12px !important;
          padding: 14px 12px 12px 12px !important;
          background:#fff !important;
          max-height: 680px !important;
          overflow-y: auto !important;
        }

        .menu-title{ font-size:22px; font-weight:800; margin:8px 0 10px 0; text-align:center; }

        div[data-testid="stImage"], div[data-testid="stImage"] * , div[data-testid="stImage"] img{
          border: 0 !important;
          outline: 0 !important;
          box-shadow: none !important;
          background: transparent !important;
          border-radius: 0 !important;
        }
        .menu-logos{ padding-top:12px !important; }
        .menu-logos, .menu-logos *{ border:0 !important; outline:0 !important; box-shadow:none !important; background:transparent !important; }
        .menu-logos div[data-testid="stVerticalBlock"],
        .menu-logos div[data-testid="stHorizontalBlock"],
        .menu-logos div { border: 0 !important; outline: 0 !important; box-shadow: none !important; }

        .sb-title{ font-size: 30px; font-weight: 900; margin: -24px 0 4px 0; }
        .sb-total{ font-size: 14px; font-weight: 700; margin: 0 0 8px 0; }

        .sb-section{
          font-size: 18px;
          font-weight: 900;
          margin: 12px 0 8px 0;
          text-align: center;
        }
        .sb-section .sb-ico{
          display: inline-block;
          margin-right: 8px;
          font-size: 18px;
          line-height: 1;
          vertical-align: middle;
        }
        .sb-subtitle{
          margin-top: 14px;
          margin-bottom: 6px;
          font-weight: 900;
        }

        .kpi-grid{ display:grid; grid-template-columns: 1fr 1fr; gap: 12px; }
        .kpi-card{
          border:1px solid #e6e6e6;
          border-radius:12px;
          padding:12px;
          background:#f9f9f9;
        }
        .kpi-title{ font-size:12px; font-weight:700; color:#111; margin-bottom:6px; }
        .kpi-value{ font-size:18px; font-weight:800; color:#000; line-height:1.1; }
        .kpi-sub{ font-size:12px; color:#333; margin-top:6px; }

        .kpi-ref{
          font-size:12px;
          font-weight:900;
          margin-top:6px;
          color:#222;
        }
        .kpi-ref .pct{
          color:#555;
          font-weight:900;
          white-space:nowrap;
        }

        div[data-baseweb="select"] > div{
          background:#fff !important;
          color:#000 !important;
          border: 1px solid #d9d9d9 !important;
          box-shadow: none !important;
        }

        div[data-testid="stMultiSelect"] span[data-baseweb="tag"]{
          background: #e8f1ff !important;
          border: 1px solid #2b6fe8 !important;
        }
        div[data-testid="stMultiSelect"] span[data-baseweb="tag"] span{
          color: #1f5fd6 !important;
          font-weight: 800 !important;
        }
        div[data-testid="stMultiSelect"] span[data-baseweb="tag"] svg{ fill: #1f5fd6 !important; }

        div[data-testid="stHorizontalBlock"]{ gap: 0.75rem !important; }
        section[data-testid="stSidebar"] hr { margin-top: 4px !important; margin-bottom: 6px !important; }

        /* =========================================================
           ✅ PATCH: HEADER/TOOLBAR BRANCO + ÍCONES OK
           (SEM mexer no menu)
        ========================================================= */
        [data-testid="stHeader"],
        [data-testid="stToolbar"],
        [data-testid="stAppToolbar"]{
          background:#ffffff !important;
          border-bottom: 1px solid #eaeaea !important;
        }

        /* Não use regra global tipo: [data-testid="stHeader"] * { ... } */
        [data-testid="stHeader"] a,
        [data-testid="stHeader"] button,
        [data-testid="stHeader"] [role="button"],
        [data-testid="stToolbar"] a,
        [data-testid="stToolbar"] button,
        [data-testid="stToolbar"] [role="button"],
        [data-testid="stAppToolbar"] a,
        [data-testid="stAppToolbar"] button,
        [data-testid="stAppToolbar"] [role="button"]{
          color:#111 !important;
        }

        /* SVGs do header: currentColor + stroke */
        [data-testid="stHeader"] svg,
        [data-testid="stToolbar"] svg,
        [data-testid="stAppToolbar"] svg{
          color:#111 !important;
          opacity: 1 !important;
        }
        [data-testid="stHeader"] svg *,
        [data-testid="stToolbar"] svg *,
        [data-testid="stAppToolbar"] svg *{
          stroke:#111 !important;
          fill: currentColor !important;
        }
        [data-testid="stHeader"] svg path[fill="none"],
        [data-testid="stToolbar"] svg path[fill="none"],
        [data-testid="stAppToolbar"] svg path[fill="none"]{
          fill: none !important;
        }

        .menu-label{
          color:#555 !important;
          font-weight:700 !important;
          opacity:1 !important;
          margin: 0 0 0.25rem 0;
          font-size: 0.85rem;
        }

        </style>
        """,
        unsafe_allow_html=True,
    )

    # reforço no head do browser (ajuda em alguns navegadores)
    st.markdown('<meta name="color-scheme" content="light">', unsafe_allow_html=True)


# =========================
# HELPERS (robustos + rápidos)
# =========================
def norm_colname(c: str) -> str:
    c = str(c).strip()
    c = re.sub(r"\s+", " ", c)
    return c.lower()


def ensure_latlon(df: pd.DataFrame, lat_candidates, lon_candidates) -> pd.DataFrame:
    df = df.copy()
    df.columns = [norm_colname(c) for c in df.columns]
    cols = set(df.columns)

    lat_col = next((c for c in lat_candidates if c in cols), None)
    lon_col = next((c for c in lon_candidates if c in cols), None)

    if lat_col is None or lon_col is None:
        raise KeyError(f"Não encontrei colunas de latitude/longitude. Disponíveis: {sorted(list(cols))[:60]}")

    df["latitude"] = df[lat_col]
    df["longitude"] = df[lon_col]
    return df


def _coerce_float(s: pd.Series) -> pd.Series:
    x = s.astype(str).str.strip()
    x = x.str.replace(r"[^0-9\-\.,]", "", regex=True)

    def _normalize_one(v: str) -> str:
        if v is None:
            return ""
        v = v.strip()
        if v.lower() in ("nan", "none", ""):
            return ""
        if v.count(",") > 1:
            parts = v.split(",")
            v = "".join(parts[:-1]) + "," + parts[-1]
        if ("," in v) and ("." in v):
            v = v.replace(".", "")
        v = v.replace(",", ".")
        return v

    x = x.apply(_normalize_one)
    x = x.replace(["nan", "none", "None", ""], np.nan)
    return pd.to_numeric(x, errors="coerce")


def _pt_number(x, nd=0):
    try:
        if x is None or (isinstance(x, (float, np.floating)) and np.isnan(x)):
            return "0"
        return f"{float(x):,.{nd}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0"


def _pct_str(part, whole, nd=0):
    try:
        part = float(part)
        whole = float(whole)
        if whole == 0:
            return "0%"
        return f"{_pt_number((part / whole) * 100, nd)}%".replace(",0%", "%")
    except Exception:
        return "0%"


def kpi_card_html(title, value, sub="Total", ref_line=None):
    ref_html = f"<div class='kpi-ref'>{ref_line}</div>" if ref_line else ""
    return (
        '<div class="kpi-card">'
        f'<div class="kpi-title">{title}</div>'
        f'<div class="kpi-value">{value}</div>'
        f'<div class="kpi-sub">{sub}</div>'
        f"{ref_html}"
        "</div>"
    )


def _safe_str(s):
    if s is None:
        return ""
    s = str(s)
    return "" if s.lower() == "nan" else s


def normaliza_dependencia(s: pd.Series) -> pd.Series:
    s = s.astype(str).replace("nan", "").str.strip()
    map_dep = {
        "1": "Federal", "01": "Federal",
        "2": "Estadual", "02": "Estadual",
        "3": "Municipal", "03": "Municipal",
        "4": "Privada", "04": "Privada"
    }
    s2 = s.replace(map_dep)
    s2 = s2.str.lower().replace({
        "federal": "Federal",
        "estadual": "Estadual",
        "municipal": "Municipal",
        "privada": "Privada",
    })
    s2 = s2.replace(map_dep)
    return s2


def escolas_por_dependencia(gdf_edu: gpd.GeoDataFrame) -> dict:
    if gdf_edu is None or gdf_edu.empty or "tp_dependencia" not in gdf_edu.columns:
        return {}
    s2 = normaliza_dependencia(gdf_edu["tp_dependencia"])
    s2 = s2[s2 != ""]
    if s2.empty:
        return {}
    return s2.value_counts().to_dict()


def escolas_por_localizacao(gdf_edu: gpd.GeoDataFrame) -> dict:
    if gdf_edu is None or gdf_edu.empty or "tp_localizacao" not in gdf_edu.columns:
        return {}

    s = gdf_edu["tp_localizacao"].astype(str).replace("nan", "").str.strip()
    s = s[s != ""]
    if s.empty:
        return {}

    map_loc = {
        "1": "Urbana", "01": "Urbana",
        "2": "Rural", "02": "Rural",
        "urbana": "Urbana",
        "rural": "Rural",
    }
    s2 = s.str.lower().replace(map_loc).replace(map_loc)
    s2 = s2.replace({"1": "Urbana", "2": "Rural", "01": "Urbana", "02": "Rural"})
    return s2.value_counts().to_dict()


def _ordered_keys(d: dict, preferred_order):
    keys = list(d.keys())
    ordered = [k for k in preferred_order if k in d]
    rest = sorted([k for k in keys if k not in ordered], key=lambda x: str(x))
    return ordered + rest


def compacto_br(n, nd=1):
    try:
        n = float(n)
    except Exception:
        return "0"
    sign = "-" if n < 0 else ""
    n = abs(n)
    if n >= 1e9:
        return f"{sign}{_pt_number(n / 1e9, nd)} Bi"
    if n >= 1e6:
        return f"{sign}{_pt_number(n / 1e6, nd)} Mi"
    if n >= 1e3:
        return f"{sign}{_pt_number(n / 1e3, nd)} Mil"
    return f"{sign}{_pt_number(n, nd)}"


def fix_mojibake_text(x: str) -> str:
    if x is None:
        return ""
    s = str(x)
    if ("Ã" in s) or ("Â" in s) or ("�" in s):
        for enc in ("latin1", "cp1252"):
            try:
                return s.encode(enc, errors="ignore").decode("utf-8", errors="ignore")
            except Exception:
                pass
    return s


def _looks_mojibake(df: pd.DataFrame) -> bool:
    try:
        obj_cols = [c for c in df.columns if df[c].dtype == "object"]
        if not obj_cols:
            return False
        sample = df[obj_cols].astype(str).head(200).values.ravel()
        txt = " ".join(sample)
        bad = txt.count("Ã") + txt.count("Â") + txt.count("�")
        return bad >= 3
    except Exception:
        return False


def read_csv_robust(path: Path, **kwargs) -> pd.DataFrame:
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin1"]
    last_err = None

    for enc in encodings:
        try:
            df = pd.read_csv(path, encoding=enc, **kwargs)

            if enc in ("cp1252", "latin1") and _looks_mojibake(df):
                for enc2 in ("utf-8-sig", "utf-8"):
                    try:
                        df2 = pd.read_csv(path, encoding=enc2, **kwargs)
                        if not _looks_mojibake(df2):
                            return df2
                    except Exception:
                        pass

            return df
        except UnicodeDecodeError as e:
            last_err = e
            continue

    if last_err:
        return pd.read_csv(path, **kwargs)
    return pd.read_csv(path, **kwargs)


# =========================
# EXPORTAÇÃO - HELPERS (DELTA)
# =========================
def _is_currency_indicator(indicador: str) -> bool:
    indicador = (indicador or "").lower()
    return ("(r$)" in indicador) or ("massa salarial" in indicador) or ("média salarial" in indicador) or (
                "media salarial" in indicador)


def _fmt_delta_cell(indicador: str, total_val, scen_val) -> str:
    """Delta = Cenário - Total, com % vs Total: '+123 (+10,5%)'."""
    if scen_val is None or (isinstance(scen_val, (float, np.floating)) and np.isnan(scen_val)):
        return ""

    try:
        sv = float(scen_val)
    except Exception:
        return ""

    try:
        tv = float(total_val) if total_val is not None else 0.0
    except Exception:
        tv = 0.0

    diff = sv - tv
    sign = "+" if diff >= 0 else "-"

    if tv == 0:
        pct_txt = "n/a"
    else:
        try:
            pct_num = (diff / tv) * 100.0
            pct_txt = f"{_pt_number(abs(pct_num), 1)}%".replace(",0%", "%")
        except Exception:
            pct_txt = "n/a"

    if _is_currency_indicator(indicador):
        diff_txt = f"R$ {_pt_number(abs(diff), 2)}"
        return f"{sign}{diff_txt} ({sign}{pct_txt})" if pct_txt != "n/a" else f"{sign}{diff_txt} (n/a)"
    else:
        if abs(diff - int(diff)) < 1e-9:
            diff_txt = _pt_number(int(abs(diff)), 0)
        else:
            diff_txt = _pt_number(abs(diff), 2)
        return f"{sign}{diff_txt} ({sign}{pct_txt})" if pct_txt != "n/a" else f"{sign}{diff_txt} (n/a)"


# =========================
# LOADERS (cacheados)
# =========================
@st.cache_data(show_spinner=False)
def load_empresas_xlsx(path: str, municipio: str) -> gpd.GeoDataFrame:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(path)

    df_raw = pd.read_excel(p)

    df = ensure_latlon(
        df_raw,
        lat_candidates=["nu_latitude", "latitude", "lat", "y"],
        lon_candidates=["nu_longitude", "longitude", "lon", "long", "x"],
    )

    if "nu_latitude" in df.columns and "nu_longitude" in df.columns:
        df["latitude"] = df["nu_latitude"]
        df["longitude"] = df["nu_longitude"]

    df["latitude"] = _coerce_float(df["latitude"])
    df["longitude"] = _coerce_float(df["longitude"])

    col_map = {
        "empregados": "Empregados",
        "massa_salarial": "Massa_Salarial",
        "média salarial": "Média Salarial",
        "media salarial": "Média Salarial",
        "cnae_2": "CNAE_2",
    }
    for k, v in col_map.items():
        if k in df.columns and v not in df.columns:
            df[v] = df[k]

    if "CNAE_2" not in df.columns:
        cnae_candidates = [c for c in df.columns if "cnae" in str(c).lower()]
        if cnae_candidates:
            c2 = [c for c in cnae_candidates if re.search(r"\b2\b", str(c)) or "2" in str(c)]
            pick = c2[0] if c2 else cnae_candidates[0]
            df["CNAE_2"] = df[pick]

    for c in ["Empregados", "Massa_Salarial", "Média Salarial"]:
        if c not in df.columns:
            df[c] = 0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    if "CNAE_2" not in df.columns:
        df["CNAE_2"] = ""

    df["Municipio"] = municipio
    df = df.dropna(subset=["latitude", "longitude"])
    df = df[(df["latitude"].between(-90, 90)) & (df["longitude"].between(-180, 180))]

    gdf = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df["longitude"], df["latitude"]), crs="EPSG:4326")
    return gdf


@st.cache_data(show_spinner=False)
def load_educacao_csv(path: str, municipio: str) -> gpd.GeoDataFrame:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(path)

    df_raw = read_csv_robust(p, sep=None, engine="python")

    df = ensure_latlon(
        df_raw,
        lat_candidates=["latitude", "lat", "nu_latitude", "nu latitude", "y"],
        lon_candidates=["longitude", "lon", "long", "nu_longitude", "nu longitude", "x"],
    )

    df["latitude"] = _coerce_float(df["latitude"])
    df["longitude"] = _coerce_float(df["longitude"])

    num_cols = [
        "qtd_prof", "qtd_matri_inf", "qtd_matri_fund", "qtd_matri_med",
        "qtd_matri_prof", "qtd_matri_eja", "qtd_matri_esp"
    ]
    for c in num_cols:
        if c not in df.columns:
            df[c] = 0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    for c in ["no_entidade", "tp_dependencia", "tp_localizacao"]:
        if c not in df.columns:
            df[c] = ""

    if "co_entidade" not in df.columns:
        df["co_entidade"] = df.index.astype(int)

    df["Municipio"] = municipio
    df = df.dropna(subset=["latitude", "longitude"])
    df = df[(df["latitude"].between(-90, 90)) & (df["longitude"].between(-180, 180))]

    gdf = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df["longitude"], df["latitude"]), crs="EPSG:4326")
    return gdf


@st.cache_data(show_spinner=False)
def load_saude_csv(path: str, municipio: str) -> gpd.GeoDataFrame:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(path)

    df_raw = read_csv_robust(p, sep=None, engine="python", dtype=str)

    df = ensure_latlon(
        df_raw,
        lat_candidates=["nu_latitude", "latitude", "lat", "y"],
        lon_candidates=["nu_longitude", "longitude", "lon", "long", "x"],
    )

    if "nu_latitude" in df.columns and "nu_longitude" in df.columns:
        df["latitude"] = df["nu_latitude"]
        df["longitude"] = df["nu_longitude"]

    df["latitude"] = _coerce_float(df["latitude"])
    df["longitude"] = _coerce_float(df["longitude"])

    for c in [
        "co_unidade", "co_cnes", "tp_unidade", "co_tipo_unidade",
        "co_tipo_estabelecimento", "no_fantasia", "no_razao_social"
    ]:
        if c not in df.columns:
            df[c] = ""

    for c in STAFF_COLS:
        if c not in df.columns:
            df[c] = 0
        df[c] = _coerce_float(df[c]).fillna(0)

    if "co_tipo_estabelecimento" in df.columns:
        df["co_tipo_estabelecimento"] = (
            df["co_tipo_estabelecimento"].astype(str).map(fix_mojibake_text).str.strip()
        )

    df["Municipio"] = municipio
    df = df.dropna(subset=["latitude", "longitude"])
    df = df[(df["latitude"].between(-90, 90)) & (df["longitude"].between(-180, 180))]

    gdf = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df["longitude"], df["latitude"]), crs="EPSG:4326")
    return gdf


@st.cache_data(show_spinner=False)
def load_all_municipios(mun_data: dict):
    emp, edu, sau = [], [], []
    for mun, paths in mun_data.items():
        emp.append(load_empresas_xlsx(paths["empresas"], mun))
        edu.append(load_educacao_csv(paths["educacao"], mun))
        sau.append(load_saude_csv(paths["saude"], mun))

    emp_all = gpd.GeoDataFrame(pd.concat(emp, ignore_index=True), geometry="geometry", crs="EPSG:4326")
    edu_all = gpd.GeoDataFrame(pd.concat(edu, ignore_index=True), geometry="geometry", crs="EPSG:4326")
    sau_all = gpd.GeoDataFrame(pd.concat(sau, ignore_index=True), geometry="geometry", crs="EPSG:4326")
    return emp_all, edu_all, sau_all


@st.cache_data(show_spinner=False)
def load_cenario_shp(shp_path: str) -> gpd.GeoDataFrame:
    gdf = gpd.read_file(shp_path)
    if gdf.crs is None:
        gdf = gdf.set_crs("EPSG:4326")
    else:
        gdf = gdf.to_crs("EPSG:4326")
    gdf = gdf.dissolve().reset_index(drop=True)
    gdf["geometry"] = gdf["geometry"].buffer(0)
    return gdf


# =========================
# MÉTRICAS
# =========================
def empresas_metrics(gdf_emp: gpd.GeoDataFrame) -> dict:
    if gdf_emp is None or gdf_emp.empty:
        return {"estab": 0, "emp": 0.0, "massa": 0.0, "media": 0.0}

    estab = int(len(gdf_emp))
    emp = float(gdf_emp["Empregados"].sum()) if "Empregados" in gdf_emp.columns else 0.0
    massa = float(gdf_emp["Massa_Salarial"].sum()) if "Massa_Salarial" in gdf_emp.columns else 0.0

    if emp > 0:
        media = massa / emp
    else:
        media = float(pd.to_numeric(gdf_emp.get("Média Salarial", pd.Series([0])), errors="coerce").fillna(0).mean())

    return {"estab": estab, "emp": emp, "massa": massa, "media": media}


def educacao_metrics(gdf_edu: gpd.GeoDataFrame) -> dict:
    if gdf_edu is None or gdf_edu.empty:
        return {
            "escolas": 0.0, "prof": 0.0,
            "inf": 0.0, "fund": 0.0, "med": 0.0,
            "profis": 0.0, "eja": 0.0, "esp": 0.0,
            "total_alunos": 0.0,
        }

    escolas = float(gdf_edu["co_entidade"].nunique()) if "co_entidade" in gdf_edu.columns else float(len(gdf_edu))
    prof = float(gdf_edu["qtd_prof"].sum()) if "qtd_prof" in gdf_edu.columns else 0.0
    inf = float(gdf_edu["qtd_matri_inf"].sum()) if "qtd_matri_inf" in gdf_edu.columns else 0.0
    fund = float(gdf_edu["qtd_matri_fund"].sum()) if "qtd_matri_fund" in gdf_edu.columns else 0.0
    med = float(gdf_edu["qtd_matri_med"].sum()) if "qtd_matri_med" in gdf_edu.columns else 0.0
    profis = float(gdf_edu["qtd_matri_prof"].sum()) if "qtd_matri_prof" in gdf_edu.columns else 0.0
    eja = float(gdf_edu["qtd_matri_eja"].sum()) if "qtd_matri_eja" in gdf_edu.columns else 0.0
    esp = float(gdf_edu["qtd_matri_esp"].sum()) if "qtd_matri_esp" in gdf_edu.columns else 0.0
    total_alunos = inf + fund + med + profis + eja + esp

    return {
        "escolas": escolas, "prof": prof,
        "inf": inf, "fund": fund, "med": med,
        "profis": profis, "eja": eja, "esp": esp,
        "total_alunos": total_alunos,
    }


def saude_metrics(gdf_s: gpd.GeoDataFrame) -> dict:
    if gdf_s is None or gdf_s.empty:
        return {"unidades": 0.0, "tipo_counts": {}, "staff_totals": {}}

    unidades = float(gdf_s["co_unidade"].nunique()) if "co_unidade" in gdf_s.columns else float(len(gdf_s))

    tipo_counts = {}
    if "co_tipo_estabelecimento" in gdf_s.columns:
        s = gdf_s["co_tipo_estabelecimento"].astype(str).map(fix_mojibake_text).str.strip()
        s = s.replace("nan", "")
        s = s[s != ""]
        if not s.empty:
            tipo_counts = s.value_counts().to_dict()

    staff_totals = {}
    for c in STAFF_COLS:
        if c in gdf_s.columns:
            staff_totals[c] = float(pd.to_numeric(gdf_s[c], errors="coerce").fillna(0).sum())
        else:
            staff_totals[c] = 0.0

    return {"unidades": unidades, "tipo_counts": tipo_counts, "staff_totals": staff_totals}


def points_in_polygon(points_gdf: gpd.GeoDataFrame, poly_gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Intersects (inclui borda). Otimizado: usa bbox + máscara vetorizada."""
    if points_gdf is None or not isinstance(points_gdf, gpd.GeoDataFrame) or points_gdf.empty:
        return gpd.GeoDataFrame(columns=getattr(points_gdf, "columns", []), geometry="geometry", crs="EPSG:4326")

    if poly_gdf is None or not isinstance(poly_gdf, gpd.GeoDataFrame) or poly_gdf.empty:
        return points_gdf.iloc[0:0].copy()

    poly = poly_gdf.geometry.iloc[0]
    try:
        poly = poly.buffer(0)
    except Exception:
        pass

    minx, miny, maxx, maxy = poly.bounds
    bbox_mask = (
            (points_gdf.geometry.x >= minx) & (points_gdf.geometry.x <= maxx) &
            (points_gdf.geometry.y >= miny) & (points_gdf.geometry.y <= maxy)
    )
    cand = points_gdf.loc[bbox_mask]
    if cand.empty:
        return cand.copy()

    mask = cand.geometry.intersects(poly)
    return cand.loc[mask].copy()


# =========================
# EXPORTAÇÃO XLSX
# =========================
def safe_sheet_name(name: str) -> str:
    n = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(name))[:31]
    return n if n.strip() else "Tabela"


def build_export_df(
        municipio: str,
        layers_sel: list,
        emp_base: dict,
        edu_base: dict,
        sau_base: dict,
        show_delta: bool,
        cenario_nome: str | None,
        emp_imp: dict | None = None,
        edu_imp: dict | None = None,
        sau_imp: dict | None = None,
        delta_layers: dict | None = None,
        gdf_edu_total: gpd.GeoDataFrame | None = None,
        gdf_edu_imp: gpd.GeoDataFrame | None = None,
        gdf_sau_total: gpd.GeoDataFrame | None = None,
        gdf_sau_imp: gpd.GeoDataFrame | None = None,
        **_ignored,
) -> pd.DataFrame:
    layers_sel = layers_sel or []
    if delta_layers is None:
        delta_layers = {
            "Empresas": ("Empresas" in layers_sel),
            "Educação": ("Educação" in layers_sel),
            "Saúde": ("Saúde" in layers_sel),
        }
    if emp_imp is None:
        emp_imp = {}
    if edu_imp is None:
        edu_imp = {}
    if sau_imp is None:
        sau_imp = {}
    if gdf_edu_total is None:
        gdf_edu_total = gpd.GeoDataFrame()
    if gdf_edu_imp is None:
        gdf_edu_imp = gpd.GeoDataFrame()
    if gdf_sau_total is None:
        gdf_sau_total = gpd.GeoDataFrame()
    if gdf_sau_imp is None:
        gdf_sau_imp = gpd.GeoDataFrame()

    scen_col = str(cenario_nome).strip() if (show_delta and cenario_nome) else None

    rows: list[dict] = []

    def _add(camada: str, indicador: str, total_val, cen_val=None):
        row = {"Município": municipio, "Camada": camada, "Indicador": indicador, "Total": total_val}
        if scen_col is not None:
            row[scen_col] = cen_val
            row["Delta"] = _fmt_delta_cell(indicador, total_val, cen_val)
        rows.append(row)

    # Empresas
    if "Empresas" in layers_sel:
        _add("Empresas", "Empregados", float(emp_base.get("emp", 0.0)),
             float(emp_imp.get("emp", np.nan)) if (show_delta and delta_layers.get("Empresas", False)) else np.nan)
        _add("Empresas", "Estabelecimentos", float(emp_base.get("estab", 0.0)),
             float(emp_imp.get("estab", np.nan)) if (show_delta and delta_layers.get("Empresas", False)) else np.nan)
        _add("Empresas", "Massa salarial (R$)", float(emp_base.get("massa", 0.0)),
             float(emp_imp.get("massa", np.nan)) if (show_delta and delta_layers.get("Empresas", False)) else np.nan)
        _add("Empresas", "Média salarial (R$)", float(emp_base.get("media", 0.0)),
             float(emp_imp.get("media", np.nan)) if (show_delta and delta_layers.get("Empresas", False)) else np.nan)

    # Saúde
    if "Saúde" in layers_sel:
        _add("Saúde", "Unidades de saúde (total)", float(sau_base.get("unidades", 0.0)),
             float(sau_imp.get("unidades", np.nan)) if (show_delta and delta_layers.get("Saúde", False)) else np.nan)

        tipo_total = {}
        if (gdf_sau_total is not None) and (not gdf_sau_total.empty) and (
                "co_tipo_estabelecimento" in gdf_sau_total.columns):
            s = gdf_sau_total["co_tipo_estabelecimento"].astype(str).map(fix_mojibake_text).str.strip()
            s = s.replace("nan", "")
            s = s[s != ""]
            if not s.empty:
                tipo_total = s.value_counts().to_dict()

        tipo_imp = {}
        if (gdf_sau_imp is not None) and (not gdf_sau_imp.empty) and ("co_tipo_estabelecimento" in gdf_sau_imp.columns):
            s = gdf_sau_imp["co_tipo_estabelecimento"].astype(str).map(fix_mojibake_text).str.strip()
            s = s.replace("nan", "")
            s = s[s != ""]
            if not s.empty:
                tipo_imp = s.value_counts().to_dict()

        if tipo_total:
            items = sorted(tipo_total.items(), key=lambda kv: kv[1], reverse=True)[:12]
            for k, v in items:
                k = fix_mojibake_text(str(k))
                _add("Saúde", f"Unidades - {k}", float(v),
                     float(tipo_imp.get(k, np.nan)) if (show_delta and delta_layers.get("Saúde", False)) else np.nan)

        staff_reg = (sau_base.get("staff_totals", {}) or {})
        staff_cen = (sau_imp.get("staff_totals", {}) or {})
        for c in STAFF_COLS:
            lab = STAFF_LABELS.get(c, c)
            _add("Saúde", f"Profissionais - {lab}", float(staff_reg.get(c, 0.0)),
                 float(staff_cen.get(c, np.nan)) if (show_delta and delta_layers.get("Saúde", False)) else np.nan)

    # Educação
    if "Educação" in layers_sel:
        _add("Educação", "Escolas", float(edu_base.get("escolas", 0.0)),
             float(edu_imp.get("escolas", np.nan)) if (show_delta and delta_layers.get("Educação", False)) else np.nan)
        _add("Educação", "Professores", float(edu_base.get("prof", 0.0)),
             float(edu_imp.get("prof", np.nan)) if (show_delta and delta_layers.get("Educação", False)) else np.nan)

        dep_total = escolas_por_dependencia(gdf_edu_total) if (
                    gdf_edu_total is not None and not gdf_edu_total.empty) else {}
        dep_imp = escolas_por_dependencia(gdf_edu_imp) if (show_delta and delta_layers.get("Educação",
                                                                                           False) and gdf_edu_imp is not None and not gdf_edu_imp.empty) else {}
        if dep_total:
            order = ["Federal", "Estadual", "Municipal", "Privada"]
            keys = _ordered_keys(dep_total, order)
            for k in keys:
                _add("Educação", f"Escolas - {k}", float(dep_total.get(k, 0)),
                     float(dep_imp.get(k, np.nan)) if (show_delta and delta_layers.get("Educação", False)) else np.nan)

        loc_total = escolas_por_localizacao(gdf_edu_total) if (
                    gdf_edu_total is not None and not gdf_edu_total.empty) else {}
        loc_imp = escolas_por_localizacao(gdf_edu_imp) if (show_delta and delta_layers.get("Educação",
                                                                                           False) and gdf_edu_imp is not None and not gdf_edu_imp.empty) else {}
        if loc_total:
            order = ["Urbana", "Rural"]
            keys = _ordered_keys(loc_total, order)
            for k in keys:
                _add("Educação", f"Escolas - {k}", float(loc_total.get(k, 0)),
                     float(loc_imp.get(k, np.nan)) if (show_delta and delta_layers.get("Educação", False)) else np.nan)

        edu_items = [
            ("Alunos - Infantil", float(edu_base.get("inf", 0.0)), float(edu_imp.get("inf", np.nan))),
            ("Alunos - Fundamental", float(edu_base.get("fund", 0.0)), float(edu_imp.get("fund", np.nan))),
            ("Alunos - Médio", float(edu_base.get("med", 0.0)), float(edu_imp.get("med", np.nan))),
            ("Alunos - Profissional", float(edu_base.get("profis", 0.0)), float(edu_imp.get("profis", np.nan))),
            ("Alunos - EJA", float(edu_base.get("eja", 0.0)), float(edu_imp.get("eja", np.nan))),
            ("Alunos - Especial", float(edu_base.get("esp", 0.0)), float(edu_imp.get("esp", np.nan))),
        ]
        for name, regv, impv in edu_items:
            _add("Educação", name, regv,
                 impv if (show_delta and delta_layers.get("Educação", False)) else np.nan)

    df = pd.DataFrame(rows)
    cam_order = {"Empresas": 1, "Saúde": 2, "Educação": 3}
    df["_cam_ord"] = df["Camada"].map(cam_order).fillna(99).astype(int)
    df = df.sort_values(["Município", "_cam_ord", "Indicador"]).drop(columns=["_cam_ord"]).reset_index(drop=True)
    return df


def export_df_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Tabela") -> bytes:
    if Workbook is None:
        raise RuntimeError("openpyxl não está instalado. Instale com: pip install openpyxl")

    if df is None or df.empty:
        df = pd.DataFrame(columns=["Município", "Camada", "Indicador", "Total"])

    value_cols = [c for c in df.columns if c not in ("Município", "Camada", "Indicador")]
    df2 = df.copy()

    numeric_cols = [c for c in value_cols if str(c).strip().lower() != "delta"]
    for c in numeric_cols:
        df2[c] = pd.to_numeric(df2[c], errors="coerce")

    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_name(sheet_name)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="111111")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for j, col in enumerate(df2.columns, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = header_align

    normal_align = Alignment(horizontal="left", vertical="center")
    num_align = Alignment(horizontal="right", vertical="center")
    currency_fmt = '"R$" #,##0.00'
    int_fmt = "#,##0"
    float_fmt = "#,##0.00"

    cols = list(df2.columns)
    idx_indicador = cols.index("Indicador") if "Indicador" in cols else None

    for i, row in enumerate(df2.itertuples(index=False, name=None), start=2):
        row_dict = dict(zip(cols, row))
        indicador = str(row[idx_indicador]).lower() if idx_indicador is not None else ""
        is_currency = ("(r$)" in indicador) or ("massa salarial" in indicador) or ("média salarial" in indicador) or (
                    "media salarial" in indicador)

        for j, col in enumerate(cols, start=1):
            val = row_dict.get(col, None)
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = border

            if col in ("Município", "Camada", "Indicador"):
                cell.alignment = normal_align
                continue

            if str(col).strip().lower() == "delta":
                cell.alignment = normal_align
                cell.number_format = "General"
                continue

            cell.alignment = num_align

            if val is None or (isinstance(val, (float, np.floating)) and np.isnan(val)):
                cell.value = None
                continue

            if isinstance(val, (np.integer, np.floating)):
                val = val.item()
                cell.value = val

            if is_currency:
                cell.number_format = currency_fmt
            else:
                try:
                    fv = float(val)
                    cell.number_format = float_fmt if abs(fv - int(fv)) > 1e-9 else int_fmt
                except Exception:
                    cell.number_format = "General"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df2.columns))}{len(df2) + 1}"

    widths = {"Município": 16, "Camada": 14, "Indicador": 34, "Total": 16, "Delta": 22}
    for j, col in enumerate(df2.columns, start=1):
        w = widths.get(col, 16 if col in value_cols else 18)
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.row_dimensions[1].height = 22

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =========================
# POPUPS
# =========================
def popup_empresas(row) -> str:
    try:
        lat = float(row.get("latitude", getattr(row.geometry, "y", np.nan)))
    except Exception:
        lat = np.nan
    try:
        lon = float(row.get("longitude", getattr(row.geometry, "x", np.nan)))
    except Exception:
        lon = np.nan

    lat_s = _pt_number(lat, 6) if not (isinstance(lat, (float, np.floating)) and np.isnan(lat)) else ""
    lon_s = _pt_number(lon, 6) if not (isinstance(lon, (float, np.floating)) and np.isnan(lon)) else ""

    return f"""
    <div style="font-family: Arial; font-size: 12px;">
      <b>Município:</b> {_safe_str(row.get('Municipio', ''))}<br>
      <b>Latitude:</b> {lat_s}<br>
      <b>Longitude:</b> {lon_s}<br>
      <b>Setor:</b> {_safe_str(row.get('CNAE_2', ''))}<br>
      <b>Nº de Empregados:</b> {int(float(row.get('Empregados', 0) or 0))}<br>
      <b>Massa Salarial:</b> R$ {_pt_number(float(row.get('Massa_Salarial', 0) or 0), 2)}<br>
    </div>
    """


def popup_educacao(row) -> str:
    try:
        lat = float(row.get("latitude", getattr(row.geometry, "y", np.nan)))
    except Exception:
        lat = np.nan
    try:
        lon = float(row.get("longitude", getattr(row.geometry, "x", np.nan)))
    except Exception:
        lon = np.nan

    lat_s = _pt_number(lat, 6) if not (isinstance(lat, (float, np.floating)) and np.isnan(lat)) else ""
    lon_s = _pt_number(lon, 6) if not (isinstance(lon, (float, np.floating)) and np.isnan(lon)) else ""

    return f"""
    <div style="font-family: Arial; font-size: 12px;">
      <b>Município:</b> {_safe_str(row.get("Municipio", ""))}<br>
      <b>Latitude:</b> {lat_s}<br>
      <b>Longitude:</b> {lon_s}<br>
      <b>Escola:</b> {_safe_str(row.get("no_entidade", ""))}<br>
      <b>Dependência:</b> {_safe_str(row.get("tp_dependencia", ""))}<br>
      <b>Localização:</b> {_safe_str(row.get("tp_localizacao", ""))}<br>
      <b>Professores:</b> {int(row.get("qtd_prof", 0))}<br><br>
      <b>Alunos por tipo</b><br>
      Infantil: {int(row.get("qtd_matri_inf", 0))}<br>
      Fundamental: {int(row.get("qtd_matri_fund", 0))}<br>
      Médio: {int(row.get("qtd_matri_med", 0))}<br>
      Profissional: {int(row.get("qtd_matri_prof", 0))}<br>
      EJA: {int(row.get("qtd_matri_eja", 0))}<br>
      Especial: {int(row.get("qtd_matri_esp", 0))}<br>
    </div>
    """


def popup_saude(row) -> str:
    try:
        lat = float(row.get("latitude", getattr(row.geometry, "y", np.nan)))
    except Exception:
        lat = np.nan
    try:
        lon = float(row.get("longitude", getattr(row.geometry, "x", np.nan)))
    except Exception:
        lon = np.nan

    lat_s = _pt_number(lat, 6) if not (isinstance(lat, (float, np.floating)) and np.isnan(lat)) else ""
    lon_s = _pt_number(lon, 6) if not (isinstance(lon, (float, np.floating)) and np.isnan(lon)) else ""

    co_cnes = _safe_str(row.get("co_cnes", ""))
    tipo = fix_mojibake_text(_safe_str(row.get("co_tipo_estabelecimento", "")))

    staff_lines = []
    any_staff = False
    for c in STAFF_COLS:
        v = row.get(c, 0)
        try:
            v = float(v)
        except Exception:
            v = 0.0
        if v > 0:
            any_staff = True
            staff_lines.append(f"{STAFF_LABELS.get(c, c)}: {int(v)}")

    staff_html = "<br>".join(staff_lines) if any_staff else "Sem informação"

    return f"""
    <div style="font-family: Arial; font-size: 12px;">
      <b>Município:</b> {fix_mojibake_text(_safe_str(row.get("Municipio", "")))}<br>
      <b>CNES:</b> {co_cnes}<br>
      <b>Tipo de Estabelecimento:</b> {tipo}<br>
      <b>Latitude:</b> {lat_s}<br>
      <b>Longitude:</b> {lon_s}<br><br>
      <b>Profissionais</b><br>
      {staff_html}
    </div>
    """


# =========================
# MAPA (otimizado)
# =========================
def build_map(center, zoom, layers_to_show, gdf_emp=None, gdf_edu=None, gdf_sau=None, cenario_poly=None):
    m = folium.Map(location=center, zoom_start=zoom, tiles="CartoDB positron", control_scale=True)

    layer_style = {
        "Empresas": {"color": "blue", "icon": "building", "prefix": "fa"},
        "Educação": {"color": "green", "icon": "graduation-cap", "prefix": "fa"},
        "Saúde": {"color": "red", "icon": "plus-square", "prefix": "fa"},
    }

    if cenario_poly is not None:
        folium.GeoJson(
            data=cenario_poly.__geo_interface__,
            name="Cenário",
            style_function=lambda x: {"fillColor": "#1f77b4", "color": "#1f77b4", "weight": 2, "fillOpacity": 0.18},
        ).add_to(m)

    def add_layer_safe(layer_name, gdf, popup_fn):
        if gdf is None or gdf.empty or (layer_name not in layers_to_show):
            return

        gdf2 = gdf[gdf.geometry.notna()].copy()
        if gdf2.empty:
            return

        stl = layer_style.get(layer_name, {"color": "blue", "icon": "info-sign", "prefix": ""})
        icon = folium.Icon(color=stl["color"], icon=stl["icon"], prefix=stl.get("prefix", ""))

        mc = MarkerCluster(name=layer_name)

        for i in range(len(gdf2)):
            geom = gdf2.geometry.iloc[i]
            if geom is None:
                continue
            row = gdf2.iloc[i]
            folium.Marker(
                location=(float(geom.y), float(geom.x)),
                popup=folium.Popup(popup_fn(row), max_width=380),
                icon=icon,
            ).add_to(mc)

        mc.add_to(m)

    add_layer_safe("Empresas", gdf_emp, popup_empresas)
    add_layer_safe("Educação", gdf_edu, popup_educacao)
    add_layer_safe("Saúde", gdf_sau, popup_saude)

    folium.LayerControl(collapsed=True).add_to(m)
    return m


# =========================
# SIDEBAR — PAINEL DE IMPACTO
# =========================
def render_impact_sidebar(
        emp_m, edu_m, sau_m,
        title_total: str,
        show_delta=False,
        base=None,
        delta_layers=None,
        gdf_edu_total=None,
        gdf_edu_imp=None,
):
    if delta_layers is None:
        delta_layers = {"Empresas": True, "Educação": True, "Saúde": True}
    if gdf_edu_total is None:
        gdf_edu_total = gpd.GeoDataFrame()
    if gdf_edu_imp is None:
        gdf_edu_imp = gpd.GeoDataFrame()

    st.sidebar.markdown('<div class="sb-title">📊 Painel de Impacto</div>', unsafe_allow_html=True)
    st.sidebar.markdown(f'<div class="sb-total">Total: {title_total}</div>', unsafe_allow_html=True)
    st.sidebar.markdown("---")

    # Empresas
    st.sidebar.markdown('<div class="sb-section"><span class="sb-ico">🏢</span>Empresas</div>', unsafe_allow_html=True)
    emp_on = bool(show_delta and base and delta_layers.get("Empresas", False))
    emp_grid = "<div class='kpi-grid'>"
    if emp_on:
        emp_grid += kpi_card_html(
            "Empresas Atingidas",
            compacto_br(emp_m["estab"], 0),
            "Total",
            ref_line=f"de {compacto_br(base['estab'], 0)} <span class='pct'>({_pct_str(emp_m['estab'], base['estab'])})</span>",
        )
        emp_grid += kpi_card_html(
            "Empregados Atingidos",
            compacto_br(emp_m["emp"], 0),
            "Total",
            ref_line=f"de {compacto_br(base['emp'], 1)} <span class='pct'>({_pct_str(emp_m['emp'], base['emp'])})</span>",
        )
        emp_grid += kpi_card_html(
            "Massa Salarial Atingida",
            f"R$ {compacto_br(emp_m['massa'], 1)}",
            "Total",
            ref_line=f"de R$ {compacto_br(base['massa'], 1)} <span class='pct'>({_pct_str(emp_m['massa'], base['massa'])})</span>",
        )
        emp_grid += kpi_card_html(
            "Média Salarial (Atingidos)",
            f"R$ {compacto_br(emp_m['media'], 1)}",
            "Média",
            ref_line=None,
        )
    else:
        emp_grid += kpi_card_html("Estabelecimentos", compacto_br(emp_m["estab"], 0), "Total")
        emp_grid += kpi_card_html("Empregados", compacto_br(emp_m["emp"], 0), "Total")
        emp_grid += kpi_card_html("Massa Salarial", f"R$ {compacto_br(emp_m['massa'], 1)}", "Total")
        emp_grid += kpi_card_html("Média Salarial", f"R$ {compacto_br(emp_m['media'], 1)}", "Média")
    emp_grid += "</div>"
    st.sidebar.markdown(emp_grid, unsafe_allow_html=True)

    # Educação
    st.sidebar.markdown('<div class="sb-section"><span class="sb-ico">🎓</span>Educação</div>', unsafe_allow_html=True)
    edu_on = bool(show_delta and base and delta_layers.get("Educação", False))

    edu_grid = "<div class='kpi-grid'>"
    if edu_on:
        edu_grid += kpi_card_html(
            "Escolas Atingidas",
            compacto_br(edu_m["escolas"], 0),
            "Total",
            ref_line=f"de {compacto_br(base['escolas'], 0)} <span class='pct'>({_pct_str(edu_m['escolas'], base['escolas'])})</span>",
        )
        edu_grid += kpi_card_html(
            "Professores Atingidos",
            compacto_br(edu_m["prof"], 0),
            "Total",
            ref_line=f"de {compacto_br(base['prof'], 0)} <span class='pct'>({_pct_str(edu_m['prof'], base['prof'])})</span>",
        )
    else:
        edu_grid += kpi_card_html("Escolas", compacto_br(edu_m["escolas"], 0), "Total")
        edu_grid += kpi_card_html("Professores", compacto_br(edu_m["prof"], 0), "Total")
    edu_grid += "</div>"
    st.sidebar.markdown(edu_grid, unsafe_allow_html=True)

    dep_total = escolas_por_dependencia(gdf_edu_total) if not gdf_edu_total.empty else {}
    dep_imp = escolas_por_dependencia(gdf_edu_imp) if (edu_on and not gdf_edu_imp.empty) else {}
    if dep_total:
        st.sidebar.markdown("<div class='sb-subtitle'>Escolas por dependência</div>", unsafe_allow_html=True)
        order = ["Federal", "Estadual", "Municipal", "Privada"]
        keys = _ordered_keys(dep_total, order)

        dep_cards = "<div class='kpi-grid'>"
        for k in keys:
            total_k = float(dep_total.get(k, 0))
            if edu_on:
                imp_k = float(dep_imp.get(k, 0))
                ref = f"de {compacto_br(total_k, 0)} <span class='pct'>({_pct_str(imp_k, total_k)})</span>"
                dep_cards += kpi_card_html(k, compacto_br(imp_k, 0), "Escolas", ref_line=ref)
            else:
                dep_cards += kpi_card_html(k, compacto_br(total_k, 0), "Escolas", ref_line=None)
        dep_cards += "</div>"
        st.sidebar.markdown(dep_cards, unsafe_allow_html=True)

    loc_total = escolas_por_localizacao(gdf_edu_total) if not gdf_edu_total.empty else {}
    loc_imp = escolas_por_localizacao(gdf_edu_imp) if (edu_on and not gdf_edu_imp.empty) else {}
    if loc_total:
        st.sidebar.markdown("<div class='sb-subtitle'>Localização</div>", unsafe_allow_html=True)
        order = ["Urbana", "Rural"]
        keys = _ordered_keys(loc_total, order)

        loc_cards = "<div class='kpi-grid'>"
        for k in keys:
            total_k = float(loc_total.get(k, 0))
            if edu_on:
                imp_k = float(loc_imp.get(k, 0))
                ref = f"de {compacto_br(total_k, 0)} <span class='pct'>({_pct_str(imp_k, total_k)})</span>"
                loc_cards += kpi_card_html(k, compacto_br(imp_k, 0), "Escolas", ref_line=ref)
            else:
                loc_cards += kpi_card_html(k, compacto_br(total_k, 0), "Escolas", ref_line=None)
        loc_cards += "</div>"
        st.sidebar.markdown(loc_cards, unsafe_allow_html=True)

    st.sidebar.markdown("<div class='sb-subtitle'>Alunos por tipo</div>", unsafe_allow_html=True)
    edu_details = [
        ("Infantil", edu_m["inf"], base.get("inf", 0) if base else 0),
        ("Fundamental", edu_m["fund"], base.get("fund", 0) if base else 0),
        ("Médio", edu_m["med"], base.get("med", 0) if base else 0),
        ("Profissional", edu_m["profis"], base.get("profis", 0) if base else 0),
        ("EJA", edu_m["eja"], base.get("eja", 0) if base else 0),
        ("Especial", edu_m["esp"], base.get("esp", 0) if base else 0),
    ]
    cards = "<div class='kpi-grid'>"
    for name, val, total_base in edu_details:
        if edu_on:
            ref = f"de {compacto_br(total_base, 0)} <span class='pct'>({_pct_str(val, total_base)})</span>"
            cards += kpi_card_html(name, compacto_br(val, 0), "Alunos", ref_line=ref)
        else:
            cards += kpi_card_html(name, compacto_br(val, 0), "Alunos", ref_line=None)
    cards += "</div>"
    st.sidebar.markdown(cards, unsafe_allow_html=True)

    # ---------- SAÚDE ----------
    st.sidebar.markdown('<div class="sb-section"><span class="sb-ico">🏥</span>Saúde</div>', unsafe_allow_html=True)
    sau_on = bool(show_delta and base and delta_layers.get("Saúde", False))

    st.sidebar.markdown("<div class='sb-subtitle'>Unidades por tipo</div>", unsafe_allow_html=True)

    tc = (sau_m.get("tipo_counts", {}) or {})
    bc = (base.get("tipo_counts", {}) if base else {})

    if tc:
        items = sorted(tc.items(), key=lambda kv: kv[1], reverse=True)[:12]
        grid = "<div class='kpi-grid'>"
        for k, v in items:
            k = fix_mojibake_text(k)
            v = float(v)
            if sau_on:
                total_k = float(bc.get(k, 0))
                ref = f"de {compacto_br(total_k, 0)} <span class='pct'>({_pct_str(v, total_k)})</span>"
                grid += kpi_card_html(f"{k}", compacto_br(v, 0), "Unidades", ref_line=ref)
            else:
                grid += kpi_card_html(f"{k}", compacto_br(v, 0), "Unidades", ref_line=None)
        grid += "</div>"
        st.sidebar.markdown(grid, unsafe_allow_html=True)
    else:
        st.sidebar.write("Sem dados de tipos.")

    st.sidebar.markdown("<div class='sb-subtitle'>Profissionais</div>", unsafe_allow_html=True)

    staff_m = (sau_m.get("staff_totals", {}) or {})
    staff_b = (base.get("staff_totals", {}) if base else {})

    staff_grid = "<div class='kpi-grid'>"
    for c in STAFF_COLS:
        label = STAFF_LABELS.get(c, c)
        val = float(staff_m.get(c, 0.0))
        if sau_on:
            total = float(staff_b.get(c, 0.0))
            ref = f"de {compacto_br(total, 0)} <span class='pct'>({_pct_str(val, total)})</span>"
            staff_grid += kpi_card_html(label, compacto_br(val, 0), "Total", ref_line=ref)
        else:
            staff_grid += kpi_card_html(label, compacto_br(val, 0), "Total", ref_line=None)
    staff_grid += "</div>"
    st.sidebar.markdown(staff_grid, unsafe_allow_html=True)


# =========================
# APP
# =========================
st.set_page_config(page_title="Impacto Econômico - RS", layout="wide", initial_sidebar_state="expanded")
inject_css()

try:
    gdf_emp_all, gdf_edu_all, gdf_sau_all = load_all_municipios(MUNICIPIOS_DATA)
except Exception as e:
    st.error(f"Erro ao carregar dados. Detalhe: {e}")
    st.info(f"Verifique se a pasta existe: {DATA_DIR}")
    st.stop()

municipios = sorted(list(MUNICIPIOS_DATA.keys()), key=lambda x: x.lower())

if "selected_mun" not in st.session_state:
    st.session_state.selected_mun = None
if "selected_cenario" not in st.session_state:
    st.session_state.selected_cenario = None

col_map, col_menu = st.columns([5.15, 1.35], gap="small")

# MENU (direita)
with col_menu:
    BID_LOGO = APP_DIR / "BID.png"
    GPEA_LOGO = APP_DIR / "GPEa.png"

    st.markdown('<div class="menu-logos">', unsafe_allow_html=True)
    cL, cR = st.columns([1, 1], gap="small")

    with cL:
        if BID_LOGO.exists():
            st.image(str(BID_LOGO), width=170)

    with cR:
        if GPEA_LOGO.exists():
            st.image(str(GPEA_LOGO), width=190)

    st.markdown("</div>", unsafe_allow_html=True)
    st.markdown('<div class="menu-title">Menu de Seleção</div>', unsafe_allow_html=True)

    st.markdown('<div class="menu-label">Selecione o Município:</div>', unsafe_allow_html=True)
    mun_choice = st.selectbox(
        "Selecione o Município:",
        options=[""] + municipios,
        index=0,
        format_func=lambda x: "Escolha um Município" if x == "" else x,
        key="mun_select",
        label_visibility="collapsed",
    )
    st.session_state.selected_mun = None if mun_choice == "" else mun_choice

    if st.session_state.selected_mun and st.session_state.selected_mun in CENARIOS:
        cen_opts = [""] + list(CENARIOS[st.session_state.selected_mun].keys())
        st.markdown('<div class="menu-label">Selecione o Cenário:</div>', unsafe_allow_html=True)
        cen_choice = st.selectbox(
            "Selecione o Cenário:",
            options=cen_opts,
            index=0,
            format_func=lambda x: "(nenhum)" if x == "" else x,
            key="cenario_select",
            label_visibility="collapsed",
        )
        st.session_state.selected_cenario = None if cen_choice == "" else cen_choice
    else:
        st.markdown('<div class="menu-label">Selecione o Cenário:</div>', unsafe_allow_html=True)
        st.selectbox(
            "Selecione o Cenário:",
            options=[""],
            format_func=lambda x: "Selecione um Município",
            disabled=True,
            key="cenario_disabled",
            label_visibility="collapsed",
        )
        st.session_state.selected_cenario = None

    try:
        st.markdown('<div class="menu-label">Exibir Camadas Atingidas:</div>', unsafe_allow_html=True)

    layers = st.multiselect(
        "Exibir Camadas Atingidas:",
        options=["Empresas", "Educação", "Saúde"],
        default=[],
        key="layers_multiselect",
        placeholder="Selecione a(s) Camada(s)",
        label_visibility="collapsed",
    )
    except TypeError:
    layers = st.multiselect(
        "Exibir Camadas Atingidas:",
        options=["Empresas", "Educação", "Saúde"],
        default=[],
        key="layers_multiselect",
        label_visibility="collapsed",
    )

if "Empresas" not in layers:
    st.session_state["filtro_setor_empresas"] = PLACEHOLDER_EMP
if "Educação" not in layers:
    st.session_state["filtro_dep_escolas"] = PLACEHOLDER_EDU
if "Saúde" not in layers:
    st.session_state["filtro_tipo_saude"] = PLACEHOLDER_SAU

if st.session_state.selected_mun:
    emp_tmp = gdf_emp_all[gdf_emp_all["Municipio"] == st.session_state.selected_mun].copy()
    edu_tmp = gdf_edu_all[gdf_edu_all["Municipio"] == st.session_state.selected_mun].copy()
    sau_tmp = gdf_sau_all[gdf_sau_all["Municipio"] == st.session_state.selected_mun].copy()
else:
    emp_tmp, edu_tmp, sau_tmp = gdf_emp_all.copy(), gdf_edu_all.copy(), gdf_sau_all.copy()

# filtro empresas
if "Empresas" in layers:
    if "CNAE_2" in emp_tmp.columns:
        setores = sorted(
            [x for x in emp_tmp["CNAE_2"].astype(str).str.strip().unique() if x and x.lower() not in ("nan", "none")])
    else:
        setores = []
    if len(setores) == 0:
        st.caption("Não foram encontrados setores (CNAE) no recorte. Verifique a coluna CNAE no arquivo.")
    st.selectbox(
        "Selecione o Setor (Empresas):",
        options=["(todos)"] + setores,
        index=0,
        key="filtro_setor_empresas",
    )
else:
    st.session_state["filtro_setor_empresas"] = "(todos)"

# filtro educação
if "Educação" in layers:
    dep_vals = normaliza_dependencia(edu_tmp.get("tp_dependencia", pd.Series([], dtype="object")))
    deps = [d for d in ["Federal", "Estadual", "Municipal", "Privada"] if d in set(dep_vals.unique())]
    st.selectbox(
        "Selecione a Dependência (Escolas):",
        options=["(todas)"] + deps,
        index=0,
        key="filtro_dep_escolas",
    )
else:
    st.session_state["filtro_dep_escolas"] = "(todas)"

# filtro saúde
if "Saúde" in layers:
    tipos = sorted([fix_mojibake_text(x) for x in
                    sau_tmp.get("co_tipo_estabelecimento", pd.Series([], dtype="object")).astype(str).unique()
                    if x and x.lower() != "nan"])
    st.selectbox(
        "Selecione a Unidade (Saúde):",
        options=["(todas)"] + tipos,
        index=0,
        key="filtro_tipo_saude",
    )
else:
    st.session_state["filtro_tipo_saude"] = "(todas)"

st.markdown("---")

layers = st.session_state.get("layers_multiselect", []) or []

# RECORTE MUNICÍPIO
if st.session_state.selected_mun:
    gdf_emp_mun = gdf_emp_all[gdf_emp_all["Municipio"] == st.session_state.selected_mun].copy()
    gdf_edu_mun = gdf_edu_all[gdf_edu_all["Municipio"] == st.session_state.selected_mun].copy()
    gdf_sau_mun = gdf_sau_all[gdf_sau_all["Municipio"] == st.session_state.selected_mun].copy()
else:
    gdf_emp_mun = gdf_emp_all.copy()
    gdf_edu_mun = gdf_edu_all.copy()
    gdf_sau_mun = gdf_sau_all.copy()

# APLICAR FILTROS
setor_sel = st.session_state.get("filtro_setor_empresas", PLACEHOLDER_EMP)
dep_sel = st.session_state.get("filtro_dep_escolas", PLACEHOLDER_EDU)
tipo_sel = st.session_state.get("filtro_tipo_saude", PLACEHOLDER_SAU)

gdf_emp_f = gdf_emp_mun.copy()
if ("Empresas" in layers) and (not is_placeholder(setor_sel, PLACEHOLDER_EMP)) and (
        str(setor_sel).strip() != "(todos)") and ("CNAE_2" in gdf_emp_f.columns):
    gdf_emp_f["CNAE_2"] = gdf_emp_f["CNAE_2"].astype(str)
    gdf_emp_f = gdf_emp_f[gdf_emp_f["CNAE_2"] == str(setor_sel)].copy()

gdf_edu_f = gdf_edu_mun.copy()
if ("Educação" in layers) and (not is_placeholder(dep_sel, PLACEHOLDER_EDU)) and (
        str(dep_sel).strip() != "(todas)") and ("tp_dependencia" in gdf_edu_f.columns):
    gdf_edu_f["_dep_norm"] = normaliza_dependencia(gdf_edu_f["tp_dependencia"])
    gdf_edu_f = gdf_edu_f[gdf_edu_f["_dep_norm"] == dep_sel].copy()

gdf_sau_f = gdf_sau_mun.copy()
if ("Saúde" in layers) and (not is_placeholder(tipo_sel, PLACEHOLDER_SAU)) and (
        str(tipo_sel).strip() != "(todas)") and ("co_tipo_estabelecimento" in gdf_sau_f.columns):
    gdf_sau_f["co_tipo_estabelecimento"] = gdf_sau_f["co_tipo_estabelecimento"].astype(str).map(
        fix_mojibake_text).str.strip()
    gdf_sau_f = gdf_sau_f[gdf_sau_f["co_tipo_estabelecimento"] == str(tipo_sel)].copy()

emp_base = empresas_metrics(gdf_emp_f)
edu_base = educacao_metrics(gdf_edu_f)
sau_base = saude_metrics(gdf_sau_f)

# CENÁRIO
cenario_gdf = None
show_delta = False
if st.session_state.selected_mun and st.session_state.selected_cenario:
    shp = CENARIOS[st.session_state.selected_mun][st.session_state.selected_cenario]
    try:
        cenario_gdf = load_cenario_shp(shp)
        show_delta = True
    except Exception as e:
        st.warning(f"Não foi possível carregar o cenário: {e}")
        cenario_gdf = None
        show_delta = False

title_total = st.session_state.selected_mun if st.session_state.selected_mun else "Lajeado + Porto Alegre + Rio Grande"

# HITS
emp_hits = edu_hits = sau_hits = None
cenario_poly = None
if show_delta and cenario_gdf is not None:
    cenario_poly = cenario_gdf.geometry.iloc[0]
    if "Empresas" in layers:
        emp_hits = points_in_polygon(gdf_emp_f, cenario_gdf)
    if "Educação" in layers:
        edu_hits = points_in_polygon(gdf_edu_f, cenario_gdf)
    if "Saúde" in layers:
        sau_hits = points_in_polygon(gdf_sau_f, cenario_gdf)

# =========================
# ✅ MAPA memoizado em session_state
# =========================
with col_map:
    if show_delta and cenario_gdf is not None:
        gdf_emp_show = emp_hits if "Empresas" in layers else None
        gdf_edu_show = edu_hits if "Educação" in layers else None
        gdf_sau_show = sau_hits if "Saúde" in layers else None
    else:
        gdf_emp_show = gdf_emp_f if "Empresas" in layers else None
        gdf_edu_show = gdf_edu_f if "Educação" in layers else None
        gdf_sau_show = gdf_sau_f if "Saúde" in layers else None

    if st.session_state.selected_mun and st.session_state.selected_mun in MUNICIPIO_VIEW:
        center = MUNICIPIO_VIEW[st.session_state.selected_mun]["center"]
        zoom = MUNICIPIO_VIEW[st.session_state.selected_mun]["zoom"]
    else:
        center, zoom = RS_CENTER, RS_ZOOM

    map_sig = (
        str(st.session_state.selected_mun or ""),
        str(st.session_state.selected_cenario or ""),
        bool(show_delta and cenario_gdf is not None),
        tuple(layers),
        str(setor_sel),
        str(dep_sel),
        str(tipo_sel),
    )

    if st.session_state.get("map_sig") != map_sig:
        st.session_state["map_sig"] = map_sig
        st.session_state["folium_map"] = build_map(
            center=center,
            zoom=zoom,
            layers_to_show=layers,
            gdf_emp=gdf_emp_show,
            gdf_edu=gdf_edu_show,
            gdf_sau=gdf_sau_show,
            cenario_poly=cenario_poly,
        )

    st_folium(
        st.session_state["folium_map"],
        height=680,
        use_container_width=True,
        key="map_main",
        returned_objects=[],
    )

st.markdown(
    """
    <div style="
        margin-top: -6px;
        margin-bottom: 4px;
        font-size: 0.75rem;
        color: #666;
        text-align: left;
    ">
        ℹ️ Os pontos no mapa foram interpolados através da API do Google com base no endereço cadastrado em cada base de dados; pontos com discrepâncias foram removidos.
    </div>
    """,
    unsafe_allow_html=True
)

# SIDEBAR (painel)
if show_delta and cenario_gdf is not None:
    emp_imp = empresas_metrics(emp_hits) if ("Empresas" in layers and emp_hits is not None) else emp_base
    edu_imp = educacao_metrics(edu_hits) if ("Educação" in layers and edu_hits is not None) else edu_base
    sau_imp = saude_metrics(sau_hits) if ("Saúde" in layers and sau_hits is not None) else sau_base

    render_impact_sidebar(
        emp_imp, edu_imp, sau_imp,
        title_total=f"{st.session_state.selected_mun} — {st.session_state.selected_cenario}",
        show_delta=True,
        base={
            "estab": emp_base["estab"], "emp": emp_base["emp"], "massa": emp_base["massa"], "media": emp_base["media"],
            "escolas": edu_base["escolas"], "prof": edu_base["prof"],
            "inf": edu_base["inf"], "fund": edu_base["fund"], "med": edu_base["med"],
            "profis": edu_base["profis"], "eja": edu_base["eja"], "esp": edu_base["esp"],
            "unidades": sau_base.get("unidades", 0.0),
            "tipo_counts": sau_base.get("tipo_counts", {}),
            "staff_totals": sau_base.get("staff_totals", {}),
        },
        delta_layers={
            "Empresas": ("Empresas" in layers),
            "Educação": ("Educação" in layers),
            "Saúde": ("Saúde" in layers),
        },
        gdf_edu_total=gdf_edu_f if ("Educação" in layers) else gpd.GeoDataFrame(),
        gdf_edu_imp=edu_hits if ("Educação" in layers and edu_hits is not None) else gpd.GeoDataFrame(),
    )
else:
    render_impact_sidebar(
        emp_base, edu_base, sau_base,
        title_total=title_total,
        show_delta=False,
        base={
            "tipo_counts": sau_base.get("tipo_counts", {}),
            "staff_totals": sau_base.get("staff_totals", {}),
        },
        delta_layers={"Empresas": False, "Educação": False, "Saúde": False},
        gdf_edu_total=gdf_edu_f if ("Educação" in layers) else gpd.GeoDataFrame(),
        gdf_edu_imp=gdf_edu_f if ("Educação" in layers) else gpd.GeoDataFrame(),
    )

# =========================
# BOTÃO DE DOWNLOAD (FINAL DO MENU) — ESTÁVEL E SEM TRAVAR
# =========================
with col_menu:
    mun_sel = st.session_state.selected_mun
    cen_sel = st.session_state.selected_cenario
    layers_sel = st.session_state.get("layers_multiselect", []) or []

    setor_sel2 = st.session_state.get("filtro_setor_empresas", "(todos)")
    dep_sel2 = st.session_state.get("filtro_dep_escolas", "(todas)")
    tipo_sel2 = st.session_state.get("filtro_tipo_saude", "(todas)")

    can_export = bool(mun_sel) and (len(layers_sel) > 0)

    if not mun_sel:
        st.info("Para exportar: selecione um Município e ao menos 1 Camada.")
    elif len(layers_sel) == 0:
        st.info("Para exportar: selecione ao menos 1 Camada em 'Exibir Camadas Atingidas'.")

    if can_export:
        has_cenario = bool(cen_sel) and bool(show_delta) and (cenario_gdf is not None)
        if (cen_sel is not None) and (str(cen_sel).strip() != "") and (not has_cenario):
            st.warning("Cenário selecionado, mas não foi possível carregá-lo. Exportando apenas o Total.")

        export_sig = (
            str(mun_sel),
            str(cen_sel) if cen_sel else "",
            bool(has_cenario),
            tuple(layers_sel),
            str(setor_sel2),
            str(dep_sel2),
            str(tipo_sel2),
        )

        sig_key = hashlib.md5(repr(export_sig).encode("utf-8")).hexdigest()

        if st.session_state.get("export_sig") != export_sig:
            st.session_state["export_sig"] = export_sig
            st.session_state["export_ready"] = False
            st.session_state["export_xlsx_bytes"] = None
            st.session_state["export_file_name"] = None

            with st.spinner("Preparando arquivo para download..."):
                emp_imp2 = empresas_metrics(emp_hits) if (
                            has_cenario and emp_hits is not None and ("Empresas" in layers_sel)) else None
                edu_imp2 = educacao_metrics(edu_hits) if (
                            has_cenario and edu_hits is not None and ("Educação" in layers_sel)) else None
                sau_imp2 = saude_metrics(sau_hits) if (
                            has_cenario and ("Saúde" in layers_sel) and sau_hits is not None) else None

                df_export = build_export_df(
                    municipio=mun_sel,
                    cenario_nome=(cen_sel if has_cenario else None),
                    layers_sel=layers_sel,
                    emp_base=emp_base, edu_base=edu_base, sau_base=sau_base,
                    emp_imp=emp_imp2, edu_imp=edu_imp2, sau_imp=sau_imp2,
                    gdf_edu_total=gdf_edu_f if ("Educação" in layers_sel) else None,
                    gdf_edu_imp=edu_hits if (
                                has_cenario and ("Educação" in layers_sel) and edu_hits is not None) else None,
                    gdf_sau_total=gdf_sau_f if ("Saúde" in layers_sel) else None,
                    gdf_sau_imp=sau_hits if (
                                has_cenario and ("Saúde" in layers_sel) and sau_hits is not None) else None,
                    show_delta=has_cenario,
                )

                layers_tag = "-".join([re.sub(r"[^A-Za-z0-9]+", "", x) for x in layers_sel])[:40]
                cen_tag = re.sub(r"[^A-Za-z0-9]+", "_", str(cen_sel)) if has_cenario else "Total"
                file_name = f"Tabela_{re.sub(r'[^A-Za-z0-9]+', '_', mun_sel)}_{cen_tag}_{layers_tag}.xlsx"

                xlsx_bytes = export_df_to_xlsx_bytes(df_export, sheet_name=mun_sel)

                st.session_state["export_xlsx_bytes"] = xlsx_bytes
                st.session_state["export_file_name"] = file_name
                st.session_state["export_ready"] = True

        if st.session_state.get("export_ready") and st.session_state.get("export_xlsx_bytes"):
            st.download_button(
                label="⬇️ Baixar Tabela (XLSX)",
                data=st.session_state["export_xlsx_bytes"],
                file_name=st.session_state["export_file_name"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"dl_{sig_key}",
            )
        else:
            st.info("Ajuste filtros/cenário/camadas para gerar o arquivo.")
